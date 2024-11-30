from datetime import datetime
import itertools
from pathlib import Path

from icalendar import Calendar, Event
from openpyxl import Workbook

from .player import Player
from .schedule import get_match_indizes_of_player, get_match_indizes_of_match
from .season import Season
from .match import convert_match_to_string, create_match


class Printer:

    def __init__(self, season: Season):
        self.season = season

    # export schedule into a excel file with each round as a row
    # consisting of the date in the first column and in every next column a match
    def export(self, folderpath: Path) -> None:
        """Export this schedule to an Excel file."""
        excel = Workbook()
        sheet = excel.active
        sheet.title = "Schedule"  # type: ignore
        sheet.append(  # type: ignore
            ["Date"] + [f"Match {i+1}" for i in range(self.season.num_courts)]
        )
        for d in sorted(self.season.dates + self.season.excluded_dates):
            if d in self.season.excluded_dates:
                sheet.append([str(d)])  # type: ignore
            else:
                i = self.season.dates.index(d)
                sheet.append([str(self.season.dates[i])] + [convert_match_to_string(m, self.season.players) for m in self.season.schedule[i]])  # type: ignore

        # add an additional sheet to excel workbook with columns for each player
        # and their match partners
        sheet = excel.create_sheet("Partner by Player")  # type: ignore
        sheet.append(  # type: ignore
            ["Date"] + [str(p) for p in self.season.players]
        )
        for i in range(len(self.season.dates)):
            row = [str(self.season.dates[i])]
            matches = self.season.schedule[i]
            for j, p in enumerate(self.season.players):
                append_string = ""
                for m in matches:
                    if j in m:
                        opponent = m[0] if m[0] != j else m[1]
                        append_string += str(self.season.players[opponent])
                        break
                row.append(append_string)
            sheet.append(row)

        # add an additional sheet to excel workbook with columns for each possible match
        # and each row marks with an x if the match is played on that day
        sheet = excel.create_sheet("Matches Overview")  # type: ignore
        player_combinations = [(p, q) for p, q in itertools.combinations(range(len(self.season.players)), 2)]
        sheet.append(  # type: ignore
            ["Date"]
            + [
                convert_match_to_string(create_match(p, q), self.season.players)
                for p, q in player_combinations
            ]
        )
        for i, round in enumerate(self.season.schedule):
            row = [str(self.season.dates[i])]
            for p, q in player_combinations:
                append_string = ""
                for m in round:
                    if m == create_match(p, q):
                        append_string = "x"
                        break
                row.append(append_string)
            sheet.append(row)

        sheet = excel.create_sheet("Costs")
        sheet.append([""] + [str(p) for p in self.season.players])  # type: ignore
        cost_per_match = self.season.overall_cost / (len(self.season.schedule) * self.season.num_courts) / 2
        sheet.append(  # type: ignore
            ["Matches"]
            + [len(get_match_indizes_of_player(self.season.schedule, p)) for p in range(len(self.season.players))]
        )
        sheet.append(  # type: ignore
            ["Cost"]
            + [len(get_match_indizes_of_player(self.season.schedule, p)) * cost_per_match for p in range(len(self.season.players))
            ]
        )

        excel.save(folderpath / "schedule.xlsx")

        # create a calendar for each player with his matches
        for i, p in enumerate(self.season.players):
            cal = Calendar()
            cal.add("prodid", "-//MatchScheduler//MatchScheduler//EN")
            cal.add("version", "2.0")
            cal.add("name", f"MatchScheduler - {p.name}")
            cal.add("X-WR-CALNAME", f"MatchScheduler - {p.name}")
            cal.add("X-WR-TIMEZONE", "Europe/Vienna")
            cal.add("X-WR-CALDESC", f"MatchScheduler - {p.name}")
            for round_index, match_index in get_match_indizes_of_player(self.season.schedule, i):
                event = Event()
                event.add("summary", self.season.calendar_title)
                event.add("description", convert_match_to_string(self.season.schedule[round_index][match_index], self.season.players))
                event.add("dtstart", datetime.combine(self.season.dates[round_index], self.season.time_start))
                event.add("dtend", datetime.combine(self.season.dates[round_index], self.season.time_end))
                cal.add_component(event)
            with open(folderpath / f"{p.name}.ics", "wb") as f:
                f.write(cal.to_ical())
