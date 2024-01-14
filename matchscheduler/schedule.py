"""a schedule class consisting of a list of rounds
a schedule is valid if all rounds are valid"""


from datetime import date, datetime, time

import numpy as np
from icalendar import Calendar, Event
from openpyxl import Workbook

from .match import Match
from .player import Player
from .round import NotValidRound, NotValidSwap, Round, RoundFactory


class Schedule:
    """A schedule of rounds."""

    def __init__(self, rounds: list[Round], players: set[Player]) -> None:
        self.rounds = rounds
        self.players = players

    def is_valid(self) -> bool:
        """Check if this schedule is valid."""
        for r in self.rounds:
            if not r.is_valid():
                return False
        return True

    def get_matches(self) -> list[Match]:
        """Get a list of all matches in this schedule."""
        return [m for r in self.rounds for m in r.matches]

    def get_matches_of_player(self, player: Player) -> list[Match]:
        """Get a list of all matches in this schedule with the given player."""
        return [m for m in self.get_matches() if player in m.players]

    def get_matches_of_players(self, players: set[Player]) -> list[Match]:
        """Get a list of all matches in this schedule with the given player."""
        return [m for m in self.get_matches() if players == m.players]

    def swap_matches(
        self, round_index1: int, match_index1: int, round_index2: int, match_index2: int
    ) -> None:
        """Swap the matches in the given rounds and match indices."""
        round1 = self.rounds[round_index1]
        round2 = self.rounds[round_index2]
        match1 = round1.matches[match_index1]
        match2 = round2.matches[match_index2]
        try:
            round1.swap_players(match_index1, match2.players)
            round2.swap_players(match_index2, match1.players)
        except NotValidSwap as exc:
            # revert swap
            round1.swap_players(match_index1, match1.players)
            round2.swap_players(match_index2, match2.players)
            # reraise error
            raise exc
        if not self.is_valid():
            raise NotValidSchedule("Swap is not valid.")

    # export schedule into a excel file with each round as a row
    # consisting of the date in the first column and in every next column a match
    def export(
        self, folderpath: str, calendar_title: str, time_start: time, time_end: time
    ) -> None:
        """Export this schedule to an Excel file."""
        excel = Workbook()
        sheet = excel.active
        sheet.title = "Schedule"  # type: ignore
        sheet.append(  # type: ignore
            ["Date"] + [f"Match {i}" for i in range(1, len(self.rounds[0].matches) + 1)]
        )
        for r in sorted(self.rounds, key=lambda r: r.date):
            sheet.append([r.date] + [str(m) for m in r.matches])  # type: ignore

        # add an additional sheet to excel workbook with columns for each player
        # and their match partners
        sheet = excel.create_sheet("Partner by Player")  # type: ignore
        sheet.append(  # type: ignore
            ["Date"] + [str(p) for p in sorted(self.players, key=lambda p: p.name)]
        )
        for r in sorted(self.rounds, key=lambda r: r.date):
            row = [str(r.date)]
            matches = r.matches
            for p in sorted(self.players, key=lambda p: p.name):
                append_string = ""
                for m in matches:
                    if p in m.players:
                        opponent = m.players.difference({p}).pop()
                        append_string = str(opponent)
                        break
                row.append(append_string)  # type: ignore
            sheet.append(row)  # type: ignore

        # add an additional sheet to excel workbook with columns for each possible match
        # and each row marks with an x if the match is played on that day
        sheet = excel.create_sheet("Matches Overview")  # type: ignore
        sheet.append(  # type: ignore
            ["Date"]
            + [Player.to_string(pm) for pm in Player.get_all_possible_combinations(self.players)]
        )
        for r in sorted(self.rounds, key=lambda r: r.date):
            row = [str(r.date)]
            matches = r.matches
            for pm in Player.get_all_possible_combinations(self.players):
                append_string = ""
                for m in matches:
                    if pm == m.players:
                        append_string = "x"
                        break
                row.append(append_string)  # type: ignore
            sheet.append(row)  # type: ignore

        excel.save(folderpath + "schedule.xlsx")

        # create a calendar for each player with his matches
        for p in self.players:
            cal = Calendar()
            cal.add("prodid", "-//MatchScheduler//MatchScheduler//EN")
            cal.add("version", "2.0")
            cal.add("name", f"MatchScheduler - {p.name}")
            cal.add("X-WR-CALNAME", f"MatchScheduler - {p.name}")
            cal.add("X-WR-TIMEZONE", "Europe/Vienna")
            cal.add("X-WR-CALDESC", f"MatchScheduler - {p.name}")
            for r in self.rounds:
                if p in r.get_players():
                    event = Event()
                    event.add("summary", calendar_title)
                    event.add("description", r.export_match_string())
                    event.add("dtstart", datetime.combine(r.date, time_start))
                    event.add("dtend", datetime.combine(r.date, time_end))
                    cal.add_component(event)
            with open(folderpath + f"{p.name}.ics", "wb") as f:
                f.write(cal.to_ical())

    def get_score(self) -> float:
        """Get the score of this schedule."""
        num_rounds = len(self.rounds)
        score = (
            num_rounds * self.get_std_of_all_possible_matches()
            + num_rounds * self.get_std_of_player_times_playing()
            + self.get_std_of_pause_between_matches()  # is calculated as sum of std
            + self.get_std_of_pause_between_playing()  # is calculated as sum of std
        )
        return score

    def get_std_of_player_times_playing(self) -> float:
        """Get the standard deviation of times playing for this schedule."""
        times_playing: dict[Player, int] = {}
        for p in self.players:
            times_playing[p] = len(self.get_matches_of_player(p))
        return np.std(list(times_playing.values()))  # type: ignore

    def get_std_of_all_possible_matches(self) -> float:
        """Get the standard deviation of all possible matches for this schedule."""
        all_possible_matches: dict[tuple[Player, Player], int] = {}
        for encounter in Player.get_all_possible_combinations(self.players):
            all_possible_matches[Player.set_to_tuple(encounter)] = len(
                self.get_matches_of_players(encounter)
            )
        return np.std(list(all_possible_matches.values()))  # type: ignore

    def get_std_of_pause_between_playing(self) -> float:
        """Get the standard deviation of pause between playing for this schedule."""
        pause_between_playing: dict[Player, float] = {}
        min_date = min((r.date for r in self.rounds))
        max_date = max((r.date for r in self.rounds))
        for p in self.players:
            matches = self.get_matches_of_player(p)
            matches.sort(key=lambda m: m.date)
            if len(matches) > 1:
                pause_between_playing[p] = np.std(  # type: ignore
                    [(matches[i + 1].date - matches[i].date).days for i in range(len(matches) - 1)]
                    + [  # add difference to first and last date
                        (matches[0].date - min_date).days,
                        (max_date - matches[-1].date).days,
                    ]
                )
            else:
                pause_between_playing[p] = (max_date - min_date).days  # set to maximal
        return np.sum(list(pause_between_playing.values()))  # type: ignore

    def get_std_of_pause_between_matches(self) -> float:
        """Get the standard deviation of pause between matches for this schedule."""
        min_date = min((r.date for r in self.rounds))
        max_date = max((r.date for r in self.rounds))
        pause_between_matches: dict[tuple[Player, Player], float] = {}
        for encounter in Player.get_all_possible_combinations(self.players):
            matches = self.get_matches_of_players(encounter)
            matches.sort(key=lambda m: m.date)
            if len(matches) > 1:
                pause_between_matches[Player.set_to_tuple(encounter)] = np.std(  # type: ignore
                    [(matches[i + 1].date - matches[i].date).days for i in range(len(matches) - 1)]
                    + [  # add difference to first and last date
                        (matches[0].date - min_date).days,
                        (max_date - matches[-1].date).days,
                    ]
                )
            else:
                pause_between_matches[Player.set_to_tuple(encounter)] = (
                    max_date - min_date  # set to maximal value
                ).days
        return np.sum(list(pause_between_matches.values()))  # type: ignore


class NotValidSchedule(Exception):
    """Raised when a schedule is not valid."""


class ScheduleFactory:
    """A factory for schedules."""

    @staticmethod
    def generate_valid_schedule(
        players: set[Player], dates: list[date], num_courts: int
    ) -> Schedule:
        """Generate a valid schedule for the given players and dates."""
        rounds = []
        for d in dates:
            try:
                rounds.append(RoundFactory.generate_valid_round(players, d, num_courts))
            except NotValidRound as exc:
                raise NotValidSchedule("Could not generate a valid schedule.") from exc

        s = Schedule(rounds, players)
        if s.is_valid():
            return s
        raise NotValidSchedule("Could not generate a valid schedule.")
